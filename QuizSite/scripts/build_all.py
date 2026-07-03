#!/usr/bin/env python3
"""
汇总题库 → manifest v3（双轨道）+ banks/

轨道 A · 考试及格：仅来自 元宇宙专业四题库.xlsx（按周次 / 按 Excel 单元列）
轨道 B · 大厂面试：成长路线面试轨
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
QUIZ = Path(__file__).resolve().parents[1]
DATA = QUIZ / "data"
BANKS = DATA / "banks"
IMPORTS = QUIZ / "imports"

INTERVIEW_MD = ROOT / "Tests" / "阶段性主程成长路线" / "大厂面试题_域映射.md"
YZ4_XLSX = ROOT / "Tests" / "元宇宙专业四题库.xlsx"

DOMAIN_NAMES = {
    1: "C# 与面向对象进阶",
    2: "Unity 核心机制深入",
    3: "渲染与图形",
    4: "物理与动画",
    5: "性能优化",
    6: "架构与工程化",
    7: "网络与多人",
    8: "编辑器扩展与工具链",
    9: "底层与引擎原理",
    10: "跨领域能力",
}
LEVEL_MAP = {"中": "中级", "高": "高级", "主程": "主程"}

CN_NUM = {
    "一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
    "六": 6, "七": 7, "八": 8, "九": 9, "十": 10,
    "十一": 11, "十二": 12, "十三": 13, "十四": 14, "十五": 15,
    "十六": 16, "十七": 17, "十八": 18, "十九": 19, "二十": 20,
}


def slug(text: str) -> str:
    s = re.sub(r"[^\w\u4e00-\u9fff]+", "-", text.strip())
    return s.strip("-").lower()[:40] or "bank"


def parse_unit_num(category: str) -> int | None:
    if not category:
        return None
    m = re.search(r"第([一二三四五六七八九十百零\d]+)单元", category)
    if not m:
        return None
    token = m.group(1)
    if token.isdigit():
        n = int(token)
        return n if 1 <= n <= 20 else None
    for key in sorted(CN_NUM.keys(), key=len, reverse=True):
        if token == key or token.startswith(key):
            return CN_NUM.get(key)
    return None


def write_bank(bank_id: str, name: str, track: str, bank_type: str, questions: list, source: str) -> dict:
    BANKS.mkdir(parents=True, exist_ok=True)
    payload = {
        "id": bank_id,
        "name": name,
        "track": track,
        "type": bank_type,
        "source": source,
        "count": len(questions),
        "questions": questions,
    }
    (BANKS / f"{bank_id}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return {
        "id": bank_id,
        "name": name,
        "count": len(questions),
        "file": f"banks/{bank_id}.json",
    }


def parse_interview() -> list:
    text = INTERVIEW_MD.read_text(encoding="utf-8")
    domains: dict[int, str] = {}
    for m in re.finditer(r"^## [三四五六七八九十]+、域 (\d+) · (.+)$", text, re.M):
        domains[int(m.group(1))] = m.group(2).strip()

    followups: dict[str, str] = {}
    for m in re.finditer(
        r"\*\*(?:中级|高级)追问答\*\*：\s*(INT-\d+-\d+)\s*→\s*「(.+?)」",
        text,
    ):
        followups[m.group(1)] = m.group(2)

    questions = []
    for line in text.splitlines():
        if not line.startswith("| INT-"):
            continue
        cols = [c.strip() for c in line.strip("|").split("|")]
        if len(cols) < 7:
            continue
        qid, freq, level_raw, qtext, focus, answer, exam = cols[:7]
        m = re.match(r"INT-(\d+)-(\d+)", qid)
        if not m:
            continue
        domain_num = int(m.group(1))
        tags = ["成长路线", "面试轨"]
        if "★" in freq:
            tags.append("高频")
        if "🔧" in freq:
            tags.append("需演示")
        if "📋" in freq:
            tags.append("偏架构")

        questions.append(
            {
                "id": qid,
                "bankId": "interview",
                "track": "interview",
                "type": "interview",
                "question": qtext,
                "answer": answer,
                "domain": domain_num,
                "domainName": domains.get(domain_num, DOMAIN_NAMES.get(domain_num, f"域{domain_num}")),
                "level": LEVEL_MAP.get(level_raw, level_raw),
                "focus": focus,
                "category": domains.get(domain_num, ""),
                "examRef": exam if exam not in ("—", "") else None,
                "tags": tags,
                "followUp": followups.get(qid),
            }
        )
    return questions


def canonical_unit_label(categories: list[str]) -> str:
    """同一单元号在 Excel 中可能写法略有差异，取出现最多的原文。"""
    if not categories:
        return ""
    return Counter(categories).most_common(1)[0][0]


def parse_exam_xlsx() -> tuple[list, dict]:
    try:
        import openpyxl
    except ImportError:
        print("  [跳过] 需 pip install openpyxl")
        return [], {}

    if not YZ4_XLSX.exists():
        print(f"  [跳过] 未找到 {YZ4_XLSX}")
        return [], {}

    wb = openpyxl.load_workbook(YZ4_XLSX, read_only=True, data_only=True)
    all_questions: list = []
    by_week: dict[int, list] = defaultdict(list)
    by_unit: dict[int, list] = defaultdict(list)
    unit_categories: dict[int, list[str]] = defaultdict(list)
    sheet_names_by_week: dict[int, str] = {}

    for i, sheet_name in enumerate(wb.sheetnames[:4], start=1):
        sheet_names_by_week[i] = sheet_name

    for sheet_name in wb.sheetnames[:4]:
        week_num = list(sheet_names_by_week.keys())[
            list(sheet_names_by_week.values()).index(sheet_name)
        ]
        ws = wb[sheet_name]
        q_idx = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2 or not row[1]:
                continue
            q_idx += 1
            num = row[0]
            qtext = str(row[1]).strip()
            answer = str(row[2] or "").strip() or "（暂无答案）"
            category = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            unit_num = parse_unit_num(category)

            if unit_num and category:
                unit_categories[unit_num].append(category)

            q = {
                "id": f"EX-W{week_num}-U{unit_num or 0:02d}-{int(num) if num else q_idx:03d}",
                "bankId": f"exam-week-{week_num}",
                "track": "exam",
                "type": "theory",
                "question": qtext,
                "answer": answer,
                "category": category,
                "unitNum": unit_num,
                "unitName": category,
                "weekNum": week_num,
                "weekName": sheet_name,
                "level": "考试及格",
                "tags": ["考试及格", sheet_name, category] if category else ["考试及格", sheet_name],
            }
            all_questions.append(q)
            by_week[week_num].append(q)
            if unit_num:
                by_unit[unit_num].append(q)

    meta: dict = {"weeks": [], "units": []}

    for w in sorted(by_week.keys()):
        qs = by_week[w]
        sheet_label = sheet_names_by_week.get(w, f"第{w}周")
        meta["weeks"].append(
            write_bank(
                f"exam-week-{w}",
                f"{sheet_label} · {len(qs)}题",
                "exam",
                "theory",
                qs,
                str(YZ4_XLSX.relative_to(ROOT)),
            )
            | {"weekNum": w, "sheetName": sheet_label}
        )

    for unit_num in sorted(by_unit.keys()):
        qs = by_unit[unit_num]
        if not qs:
            continue
        label = canonical_unit_label(unit_categories[unit_num])
        meta["units"].append(
            write_bank(
                f"exam-unit-u{unit_num:02d}",
                f"{label} · {len(qs)}题",
                "exam",
                "theory",
                qs,
                str(YZ4_XLSX.relative_to(ROOT)),
            )
            | {"unitNum": unit_num, "unitName": label}
        )

    write_bank(
        "exam-all",
        f"全部理论题 · {len(all_questions)}题",
        "exam",
        "theory",
        all_questions,
        str(YZ4_XLSX.relative_to(ROOT)),
    )

    print(
        f"  考试理论: {len(all_questions)} 题 · 周次 {len(meta['weeks'])} 卷"
        f" · 单元 {len(meta['units'])} 个（仅 Excel 已有）"
    )
    for u in meta["units"]:
        print(f"    {u['unitName']}: {u['count']} 题")

    return all_questions, meta


def parse_imports() -> list[dict]:
    IMPORTS.mkdir(parents=True, exist_ok=True)
    out = []
    for path in sorted(IMPORTS.iterdir()):
        if path.name.lower().startswith("template"):
            continue
        if path.suffix.lower() == ".json":
            data = json.loads(path.read_text(encoding="utf-8"))
            questions = data if isinstance(data, list) else data.get("questions", [])
            name = data.get("name", path.stem) if isinstance(data, dict) else path.stem
        elif path.suffix.lower() == ".csv":
            questions = []
            with path.open(encoding="utf-8-sig", newline="") as f:
                for i, row in enumerate(csv.DictReader(f)):
                    q = row.get("question") or row.get("题干") or row.get("题目") or ""
                    if not q.strip():
                        continue
                    questions.append(
                        {
                            "question": q.strip(),
                            "answer": (row.get("answer") or row.get("答案") or "").strip(),
                            "category": (row.get("category") or row.get("单元") or "").strip(),
                        }
                    )
            name = path.stem
        else:
            continue
        if not questions:
            continue
        bank_id = f"custom-{slug(path.stem)}"
        normalized = []
        for i, q in enumerate(questions):
            normalized.append(
                {
                    "id": q.get("id") or f"CUS-{i+1:04d}",
                    "bankId": bank_id,
                    "track": "exam",
                    "type": "custom",
                    "question": q.get("question", ""),
                    "answer": q.get("answer", ""),
                    "category": q.get("category", ""),
                    "level": "考试及格",
                    "tags": ["自定义"],
                }
            )
        out.append(
            write_bank(bank_id, name, "exam", "custom", normalized, f"imports/{path.name}")
        )
        print(f"  自定义: {path.name} → {len(normalized)} 题")
    return out


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    BANKS.mkdir(parents=True, exist_ok=True)

    print("构建大厂面试 · 成长路线…")
    interview_q = parse_interview()
    interview_bank = write_bank(
        "interview",
        "大厂面试 · 成长路线",
        "interview",
        "interview",
        interview_q,
        "Tests/阶段性主程成长路线/大厂面试题_域映射.md",
    )
    print(f"  面试题: {len(interview_q)} 题")

    print("构建考试及格 · 理论题库…")
    _, exam_meta = parse_exam_xlsx()
    custom = parse_imports()

    manifest = {
        "version": 3,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "tracks": [
            {
                "id": "exam",
                "name": "考试及格 · 理论题库",
                "subtitle": "元宇宙专业四题库.xlsx · 与面试轨分开",
                "defaultMode": "week",
                "defaultSelection": "exam-week-1",
                "modes": {
                    "week": {
                        "label": "按周次",
                        "hint": "对应 Excel 四个工作表，每卷约 64 题",
                        "options": exam_meta.get("weeks", []),
                    },
                    "unit": {
                        "label": "按单元",
                        "hint": "仅列出 Excel「对应单元」列中实际有题的单元",
                        "options": exam_meta.get("units", []),
                    },
                },
                "customBanks": custom,
            },
            {
                "id": "interview",
                "name": "大厂面试 · 成长路线",
                "subtitle": "口述面试轨 · 不进考试及格线",
                "file": interview_bank["file"],
                "count": interview_bank["count"],
                "bankId": "interview",
            },
        ],
    }

    keep_ids = {"interview", "exam-all"}
    keep_ids.update(o["id"] for o in exam_meta.get("weeks", []))
    keep_ids.update(o["id"] for o in exam_meta.get("units", []))
    keep_ids.update(c["id"] for c in custom)
    for old in BANKS.glob("*.json"):
        if old.stem not in keep_ids and (old.stem.startswith("yz4-") or old.stem.startswith("exam-")):
            old.unlink()

    (DATA / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"\n完成 → {DATA / 'manifest.json'}")


if __name__ == "__main__":
    main()
